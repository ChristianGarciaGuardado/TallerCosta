from flask import Flask, render_template, request, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import io

app = Flask(__name__, static_folder='static', static_url_path='/static')

# ─── Base de datos ────────────────────────────────────────────────────────────
# Usa PostgreSQL en Render (variable DATABASE_URL) o SQLite localmente
DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///taller.db')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'taller_costa_2025'

db = SQLAlchemy(app)

# ═══════════════════════════════════════════════════════
# MODELOS — Tablas de la base de datos
# ═══════════════════════════════════════════════════════

class Cliente(db.Model):
    """Clientes del taller"""
    id = db.Column(db.Integer, primary_key=True)
    empresa = db.Column(db.String(100), nullable=False)
    contacto = db.Column(db.String(100))
    telefono = db.Column(db.String(30))
    email = db.Column(db.String(100))
    direccion = db.Column(db.String(200))
    cuit = db.Column(db.String(20))
    notas = db.Column(db.Text)
    creado = db.Column(db.DateTime, default=datetime.utcnow)
    presupuestos = db.relationship('Presupuesto', backref='cliente', lazy=True)
    trabajos = db.relationship('Trabajo', backref='cliente', lazy=True)

class Presupuesto(db.Model):
    """Presupuestos generados para clientes"""
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(20), unique=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    tipo_equipo = db.Column(db.String(50))
    identificador = db.Column(db.String(50))
    marca = db.Column(db.String(50))
    modelo = db.Column(db.String(50))
    tipo_trabajo = db.Column(db.String(50))
    observaciones = db.Column(db.Text)
    estado = db.Column(db.String(20), default='borrador')  # borrador, enviado, aceptado, rechazado
    total = db.Column(db.Float, default=0)
    creado = db.Column(db.DateTime, default=datetime.utcnow)
    items = db.relationship('ItemPresupuesto', backref='presupuesto', lazy=True, cascade='all, delete-orphan')
    trabajo = db.relationship('Trabajo', backref='presupuesto', lazy=True, uselist=False)

class ItemPresupuesto(db.Model):
    """Ítems de cada presupuesto (descripcion, cantidad, precio)"""
    id = db.Column(db.Integer, primary_key=True)
    presupuesto_id = db.Column(db.Integer, db.ForeignKey('presupuesto.id'), nullable=False)
    descripcion = db.Column(db.String(200), nullable=False)
    cantidad = db.Column(db.Float, default=1)
    precio_unitario = db.Column(db.Float, default=0)
    descuento = db.Column(db.Float, default=0)  # porcentaje 0-100
    subtotal = db.Column(db.Float, default=0)

class Trabajo(db.Model):
    """Trabajos activos e históricos del taller"""
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(20), unique=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    presupuesto_id = db.Column(db.Integer, db.ForeignKey('presupuesto.id'), nullable=True)
    tipo_equipo = db.Column(db.String(50))
    identificador = db.Column(db.String(50))
    marca = db.Column(db.String(50))
    modelo = db.Column(db.String(50))
    tipo_trabajo = db.Column(db.String(50))
    observaciones = db.Column(db.Text)
    estado = db.Column(db.String(20), default='en_curso')  # en_curso, finalizado, entregado
    presupuestado = db.Column(db.Float, default=0)
    fecha_ingreso = db.Column(db.Date, default=date.today)
    fecha_entrega = db.Column(db.Date, nullable=True)
    creado = db.Column(db.DateTime, default=datetime.utcnow)
    gastos = db.relationship('GastoTrabajo', backref='trabajo', lazy=True, cascade='all, delete-orphan')
    cobros = db.relationship('Cobro', backref='trabajo', lazy=True, cascade='all, delete-orphan')

    @property
    def total_gastos(self):
        return sum(g.monto for g in self.gastos)

    @property
    def total_cobrado(self):
        return sum(c.monto for c in self.cobros)

    @property
    def saldo(self):
        return self.presupuestado - self.total_cobrado

    @property
    def ganancia(self):
        return self.total_cobrado - self.total_gastos

class GastoTrabajo(db.Model):
    """Gastos asociados a un trabajo específico (repuestos, insumos, etc.)"""
    id = db.Column(db.Integer, primary_key=True)
    trabajo_id = db.Column(db.Integer, db.ForeignKey('trabajo.id'), nullable=False)
    fecha = db.Column(db.Date, default=date.today)
    categoria = db.Column(db.String(50))
    proveedor = db.Column(db.String(100))
    descripcion = db.Column(db.String(200))
    monto = db.Column(db.Float, default=0)
    forma_pago = db.Column(db.String(30))

class Cobro(db.Model):
    """Pagos recibidos por un trabajo"""
    id = db.Column(db.Integer, primary_key=True)
    trabajo_id = db.Column(db.Integer, db.ForeignKey('trabajo.id'), nullable=False)
    fecha = db.Column(db.Date, default=date.today)
    monto = db.Column(db.Float, default=0)
    forma_pago = db.Column(db.String(30))
    comprobante = db.Column(db.String(50))
    notas = db.Column(db.Text)

class GastoGeneral(db.Model):
    """Gastos del taller no asociados a un trabajo (alquiler, servicios, etc.)"""
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, default=date.today)
    categoria = db.Column(db.String(50))
    proveedor = db.Column(db.String(100))
    monto = db.Column(db.Float, default=0)
    forma_pago = db.Column(db.String(30))
    comprobante = db.Column(db.String(50))
    anulado = db.Column(db.Boolean, default=False)
    notas = db.Column(db.Text)

class OpcionLista(db.Model):
    """Opciones configurables de los desplegables (tipos de equipo, trabajo, etc.)"""
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(50), nullable=False)   # tipo_equipo, tipo_trabajo, categoria_gasto, forma_pago
    valor = db.Column(db.String(100), nullable=False)
    orden = db.Column(db.Integer, default=0)          # para ordenar los items en el desplegable

    def __repr__(self):
        return f'<OpcionLista {self.tipo}: {self.valor}>'

# ═══════════════════════════════════════════════════════
# HELPERS — Funciones auxiliares
# ═══════════════════════════════════════════════════════

def gen_numero_presupuesto():
    """Genera el próximo número de presupuesto (PRES-0001, PRES-0002, etc.)"""
    ultimo = Presupuesto.query.order_by(Presupuesto.id.desc()).first()
    n = (ultimo.id + 1) if ultimo else 1
    return f"PRES-{n:04d}"

def gen_numero_trabajo():
    """Genera el próximo número de reparación (REP-0001, REP-0002, etc.)"""
    ultimo = Trabajo.query.order_by(Trabajo.id.desc()).first()
    n = (ultimo.id + 1) if ultimo else 1
    return f"REP-{n:04d}"

def fmt_moneda(valor):
    """Formatea un número como moneda argentina ($1.234.567)"""
    if valor is None:
        return "$0"
    return f"${valor:,.0f}".replace(",", ".")

def get_opciones(tipo):
    """Devuelve la lista de opciones para un tipo de desplegable"""
    return [o.valor for o in OpcionLista.query.filter_by(tipo=tipo).order_by(OpcionLista.valor).all()]

def seed_opciones():
    """Carga las opciones por defecto si la tabla está vacía"""
    if OpcionLista.query.count() == 0:
        defaults = {
            'tipo_equipo': ['Camión', 'Pluma / Grúa', 'Autoelevador / Clark',
                            'Excavadora', 'Retroexcavadora', 'Compactadora',
                            'Generador', 'Compresor', 'Otro'],
            'tipo_trabajo': ['Mecánica general', 'Motor', 'Transmisión',
                             'Hidráulica', 'Eléctrico', 'Neumático',
                             'Chapa y pintura', 'Restauración', 'Preventivo', 'Otro'],
            'categoria_gasto': ['Repuestos', 'Insumos', 'Mano de Obra Externa',
                                'Herramientas', 'Alquiler', 'Servicios (luz/agua/gas)',
                                'Seguros', 'Impuestos', 'Administración', 'Fletes', 'Otros'],
            'forma_pago': ['Efectivo', 'Transferencia', 'Cheque',
                           'Tarjeta Débito', 'Tarjeta Crédito', 'Cuenta Corriente'],
        }
        for tipo, valores in defaults.items():
            for i, valor in enumerate(valores):
                db.session.add(OpcionLista(tipo=tipo, valor=valor, orden=i))
        db.session.commit()

app.jinja_env.filters['moneda'] = fmt_moneda

# ═══════════════════════════════════════════════════════
# RUTAS — INICIO
# ═══════════════════════════════════════════════════════

@app.route('/')
def inicio():
    """Pantalla principal con KPIs comerciales por rango de fechas y listados operativos"""
    hoy = date.today()
    
    # Filtro por rango de fechas (por defecto: desde el día 1 de este mes hasta hoy)
    fecha_desde_str = request.args.get('desde', '')
    fecha_hasta_str = request.args.get('hasta', '')
    
    if fecha_desde_str:
        fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
    else:
        fecha_desde = date(hoy.year, hoy.month, 1)
        fecha_desde_str = fecha_desde.strftime('%Y-%m-%d')
        
    if fecha_hasta_str:
        fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
    else:
        fecha_hasta = hoy
        fecha_hasta_str = fecha_hasta.strftime('%Y-%m-%d')

    # 1. LISTADO DE TRABAJOS ACTIVOS (Fijos del taller hoy, excluyendo 'entregado' y 'anulado')
    trabajos_activos = Trabajo.query.filter(
        Trabajo.estado.in_(['en_curso', 'finalizado'])
    ).order_by(Trabajo.creado.desc()).all()

    # 2. LISTADO DE PRESUPUESTOS PENDIENTES (Solo los que esperan aprobación)
    presupuestos_pendientes_list = Presupuesto.query.filter(
        Presupuesto.estado.in_(['borrador', 'enviado', 'pendiente'])
    ).order_by(Presupuesto.creado.desc()).all()

    # 3. CÁLCULO DE KPIs
    # KPI 1: Cantidad de presupuestos pendientes
    cant_presupuestos = len(presupuestos_pendientes_list)
    
    # KPI 2: Suma de dinero potencial de lo que está pendiente en la calle
    total_dinero_pendientes = sum(p.total for p in presupuestos_pendientes_list)

    # KPI 3: Ventas del período (Suma directa del campo 'presupuestado' de los trabajos creados en el rango, sin anulados)
    trabajos_del_periodo = Trabajo.query.filter(
        Trabajo.creado >= fecha_desde,
        Trabajo.creado <= fecha_hasta,
        Trabajo.estado != 'anulado'
    ).all()
    ventas_del_periodo = sum(t.presupuestado for t in trabajos_del_periodo)

    # KPI 4: Saldo total histórico a cobrar (De los trabajos activos en este momento)
    total_saldo = sum(t.saldo for t in trabajos_activos)

    return render_template('inicio.html',
                           trabajos=trabajos_activos,
                           presupuestos_lista=presupuestos_pendientes_list,
                           presupuestos_pendientes=cant_presupuestos,
                           total_dinero_pendientes=total_dinero_pendientes,
                           ventas_este_mes=ventas_del_periodo,
                           total_saldo=total_saldo,
                           fecha_desde=fecha_desde_str,
                           fecha_hasta=fecha_hasta_str)

# ═══════════════════════════════════════════════════════
# RUTAS — CLIENTES
# ═══════════════════════════════════════════════════════

@app.route('/clientes')
def clientes():
    q = request.args.get('q', '')
    if q:
        lista = Cliente.query.filter(Cliente.empresa.ilike(f'%{q}%')).order_by(Cliente.empresa).all()
    else:
        lista = Cliente.query.order_by(Cliente.empresa).all()
    return render_template('clientes.html', clientes=lista, q=q)

@app.route('/clientes/nuevo', methods=['GET', 'POST'])
def nuevo_cliente():
    if request.method == 'POST':
        c = Cliente(
            empresa=request.form['empresa'],
            contacto=request.form.get('contacto'),
            telefono=request.form.get('telefono'),
            email=request.form.get('email'),
            direccion=request.form.get('direccion'),
            cuit=request.form.get('cuit'),
            notas=request.form.get('notas')
        )
        db.session.add(c)
        db.session.commit()
        return redirect(url_for('clientes'))
    return render_template('cliente_form.html', cliente=None)

@app.route('/clientes/<int:id>/editar', methods=['GET', 'POST'])
def editar_cliente(id):
    c = Cliente.query.get_or_404(id)
    if request.method == 'POST':
        c.empresa = request.form['empresa']
        c.contacto = request.form.get('contacto')
        c.telefono = request.form.get('telefono')
        c.email = request.form.get('email')
        c.direccion = request.form.get('direccion')
        c.cuit = request.form.get('cuit')
        c.notas = request.form.get('notas')
        db.session.commit()
        return redirect(url_for('clientes'))
    return render_template('cliente_form.html', cliente=c)

# ═══════════════════════════════════════════════════════════════════════
# RUTAS — PRESUPUESTOS (CONSOLIDADO Y SEGURO)
# ═══════════════════════════════════════════════════════════════════════

@app.route('/presupuestos')
def presupuestos():
    """Listado general de presupuestos filtrado por pestañas en la URL"""
    estado_filtro = request.args.get('estado', 'activos')
    
    # 1. Determinamos qué mostrar según la pestaña seleccionada
    if estado_filtro == 'aceptados':
        lista_presupuestos = Presupuesto.query.filter_by(estado='aceptado').order_by(Presupuesto.creado.desc()).all()
    elif estado_filtro == 'rechazados':
        lista_presupuestos = Presupuesto.query.filter_by(estado='rechazado').order_by(Presupuesto.creado.desc()).all()
    else:  # 'activos' -> Muestra borradores y enviados
        lista_presupuestos = Presupuesto.query.filter(Presupuesto.estado.in_(['borrador', 'enviado'])).order_by(Presupuesto.creado.desc()).all()
    
    # 2. Obtenemos los IDs de presupuestos cuyos trabajos siguen activos en el taller para habilitar el botón "Ampliar"
    trabajos_activos = Trabajo.query.filter(Trabajo.estado.in_(['en_curso', 'finalizado'])).all()
    ids_presupuestos_activos = [t.presupuesto_id for t in trabajos_activos if t.presupuesto_id]
    
    return render_template('presupuestos.html', 
                           lista=lista_presupuestos, 
                           estado=estado_filtro, 
                           ids_activos=ids_presupuestos_activos)


@app.route('/presupuestos/<int:id>/editar', methods=['GET', 'POST'])
def editar_presupuesto(id):
    """Edición inteligente de presupuestos pendientes o ampliaciones en taller"""
    p = Presupuesto.query.get_or_404(id)
    
    # ════════════════════════════════════════════════════════════════════
    # VALIDADOR INTELIGENTE DE EDICIÓN
    # ════════════════════════════════════════════════════════════════════
    if p.estado == 'rechazado':
        return "Este presupuesto fue rechazado y no se puede modificar.", 403
        
    if p.estado == 'aceptado':
        # Buscamos si el trabajo asociado a este presupuesto sigue activo en el taller
        trabajo_asociado = Trabajo.query.filter(
            Trabajo.presupuesto_id == p.id,
            Trabajo.estado.in_(['en_curso', 'finalizado'])
        ).first()
        
        # Validamos el resultado para que VS Code use la variable y no quede opaca
        if not trabajo_asociado:
            return "No se puede editar este presupuesto porque el trabajo ya finalizó, fue entregado o anulado.", 403
    # ════════════════════════════════════════════════════════════════════

    clientes = Cliente.query.order_by(Cliente.empresa).all()
    tipos_equipo = ['Camión', 'Utilitario', 'Acoplado', 'Máquina Vial']
    tipos_trabajo = ['Mecánica General', 'Electricidad', 'Frenos', 'Motor', 'Service']
    
    if request.method == 'POST':
        p.cliente_id = int(request.form.get('cliente_id'))
        p.tipo_equipo = request.form.get('tipo_equipo')
        p.marca = request.form.get('marca')
        p.modelo = request.form.get('modelo')
        p.identificador = request.form.get('identificador')
        p.tipo_trabajo = request.form.get('tipo_trabajo')
        p.observaciones = request.form.get('observaciones')
        
        # Si el presupuesto NO estaba aceptado, cambia su estado según el botón presionado.
        # Si YA estaba aceptado, preserva el estado 'aceptado'.
        if p.estado not in ['aceptado']:
            p.estado = request.form.get('estado', 'borrador')
        
        # Procesamos ítems cuidando la consistencia de nombres del HTML
        descripciones = request.form.getlist('descripcion[]')
        cantidades = request.form.getlist('cantidad[]')
        precios = request.form.getlist('precio[]')  # Sincronizado con name="precio[]"
        descuentos = request.form.getlist('descuento[]')
        
        # Eliminamos ítems viejos para reescribir
        for item in p.items:
            db.session.delete(item)
            
        total_general = 0
        for i in range(len(descripciones)):
            if descripciones[i].strip():
                cant = float(cantidades[i]) if (i < len(cantidades) and cantidades[i]) else 1.0
                precio = float(precios[i]) if (i < len(precios) and precios[i]) else 0.0
                desc = float(descuentos[i]) if (i < len(descuentos) and descuentos[i]) else 0.0
                
                subtotal = cant * precio * (1 - desc / 100)
                total_general += subtotal
                
                nuevo_item = ItemPresupuesto(
                    presupuesto_id=p.id,
                    descripcion=descripciones[i],
                    cantidad=cant,
                    precio_unitario=precio,
                    descuento=desc,
                    subtotal=subtotal
                )
                db.session.add(nuevo_item)
        
        p.total = total_general
        
        # Si es una ampliación, sincronizamos el monto de forma automática en el taller
        if p.estado == 'aceptado':
            trabajo_taller = Trabajo.query.filter_by(presupuesto_id=p.id).first()
            if trabajo_taller:
                trabajo_taller.presupuestado = total_general
                trabajo_taller.marca = p.marca
                trabajo_taller.modelo = p.modelo
                trabajo_taller.identificador = p.identificador
                trabajo_taller.tipo_trabajo = p.tipo_trabajo
        
        db.session.commit()
        return redirect(url_for('presupuestos'))
        
    return render_template('presupuesto_form.html', 
                           presupuesto=p, 
                           clientes=clientes, 
                           tipos_equipo=tipos_equipo, 
                           tipos_trabajo=tipos_trabajo)


@app.route('/presupuestos/<int:id>/estado', methods=['POST'])
def cambiar_estado_presupuesto(id):
    """Control de cambios de estado directos desde los botones del listado"""
    p = Presupuesto.query.get_or_404(id)
    nuevo_estado = request.form.get('estado')
    
    p.estado = nuevo_estado
    db.session.commit()
    return redirect(url_for('presupuestos'))

# ═══════════════════════════════════════════════════════
# RUTAS — TRABAJOS
# ═══════════════════════════════════════════════════════

@app.route('/trabajos')
def trabajos():
    cliente_id = request.args.get('cliente_id', '')
    estado = request.args.get('estado', '')
    maquina = request.args.get('maquina', '')
    q = Trabajo.query
    if cliente_id:
        q = q.filter_by(cliente_id=cliente_id)
    if estado:
        q = q.filter_by(estado=estado)
    if maquina:
        q = q.filter(Trabajo.identificador.ilike(f'%{maquina}%'))
    lista = q.order_by(Trabajo.creado.desc()).all()
    clientes = Cliente.query.order_by(Cliente.empresa).all()
    return render_template('trabajos.html', trabajos=lista, clientes=clientes,
                           cliente_id=cliente_id, estado=estado, maquina=maquina)

@app.route('/trabajos/<int:id>')
def detalle_trabajo(id):
    t = Trabajo.query.get_or_404(id)
    categorias = get_opciones('categoria_gasto')
    formas_pago = get_opciones('forma_pago')
    return render_template('detalle_trabajo.html', trabajo=t,
                           categorias=categorias, formas_pago=formas_pago)

@app.route('/trabajos/<int:id>/estado', methods=['POST'])
def cambiar_estado_trabajo(id):
    t = Trabajo.query.get_or_404(id)
    t.estado = request.form['estado']
    if t.estado == 'entregado':
        t.fecha_entrega = date.today()
    db.session.commit()
    return redirect(url_for('detalle_trabajo', id=id))

@app.route('/trabajos/<int:id>/gasto', methods=['POST'])
def agregar_gasto_trabajo(id):
    Trabajo.query.get_or_404(id)
    g = GastoTrabajo(
        trabajo_id=id,
        fecha=datetime.strptime(request.form['fecha'], '%Y-%m-%d').date(),
        categoria=request.form['categoria'],
        proveedor=request.form.get('proveedor'),
        descripcion=request.form.get('descripcion'),
        monto=float(request.form['monto'] or 0),
        forma_pago=request.form.get('forma_pago')
    )
    db.session.add(g)
    db.session.commit()
    return redirect(url_for('detalle_trabajo', id=id))

@app.route('/trabajos/gasto/<int:id>/eliminar', methods=['POST'])
def eliminar_gasto_trabajo(id):
    g = GastoTrabajo.query.get_or_404(id)
    trabajo_id = g.trabajo_id
    db.session.delete(g)
    db.session.commit()
    return redirect(url_for('detalle_trabajo', id=trabajo_id))

@app.route('/trabajos/<int:id>/cobro', methods=['POST'])
def agregar_cobro(id):
    t = Trabajo.query.get_or_404(id)
    c = Cobro(
        trabajo_id=id,
        fecha=datetime.strptime(request.form['fecha'], '%Y-%m-%d').date(),
        monto=float(request.form['monto'] or 0),
        forma_pago=request.form['forma_pago'],
        comprobante=request.form.get('comprobante'),
        notas=request.form.get('notas')
    )
    db.session.add(c)
    db.session.commit()
    return redirect(url_for('cuenta_corriente') + f'?cliente_id={t.cliente_id}')

@app.route('/trabajos/cobro/<int:id>/eliminar', methods=['POST'])
def eliminar_cobro(id):
    c = Cobro.query.get_or_404(id)
    t = Trabajo.query.get(c.trabajo_id)
    db.session.delete(c)
    db.session.commit()
    return redirect(url_for('cuenta_corriente') + f'?cliente_id={t.cliente_id}')

@app.route('/trabajos/<int:id>/notas', methods=['POST'])
def actualizar_notas_trabajo(id):
    """Actualiza las observaciones/detalle de un trabajo"""
    t = Trabajo.query.get_or_404(id)
    t.observaciones = request.form.get('observaciones', '')
    db.session.commit()
    return redirect(url_for('detalle_trabajo', id=id))

# ═══════════════════════════════════════════════════════
# RUTAS — CUENTA CORRIENTE
# ═══════════════════════════════════════════════════════

@app.route('/cuenta-corriente')
def cuenta_corriente():
    cliente_id = request.args.get('cliente_id', '')
    clientes = Cliente.query.order_by(Cliente.empresa).all()
    formas_pago = get_opciones('forma_pago')
    cliente_sel = None
    trabajos = []
    if cliente_id:
        cliente_sel = Cliente.query.get(cliente_id)
        trabajos = Trabajo.query.filter_by(cliente_id=cliente_id).order_by(Trabajo.creado.desc()).all()
    return render_template('cuenta_corriente.html',
        clientes=clientes, cliente_sel=cliente_sel,
        trabajos=trabajos, cliente_id=cliente_id, formas_pago=formas_pago)

# ═══════════════════════════════════════════════════════
# RUTAS — HISTORIAL
# ═══════════════════════════════════════════════════════

@app.route('/historial')
def historial():
    """Historial técnico global: búsquedas exactas por desplegables (activos + pasados)"""
    cliente_q = request.args.get('cliente', '')
    maquina_q = request.args.get('maquina', '')
    
    # Consulta limpia para traer todos los estados (en curso, finalizados, entregados)
    q = Trabajo.query
    
    # CAMBIO: Comparación exacta (==) para el nombre del cliente seleccionado
    if cliente_q:
        q = q.join(Cliente).filter(Cliente.empresa == cliente_q)
        
    # CAMBIO: Comparación exacta (==) para evitar cruces si hay patentes/internos parecidos
    if maquina_q:
        q = q.filter(Trabajo.identificador == maquina_q)
        
    # Ordenamos por ingreso: lo último que entró al taller figura arriba de todo
    lista = q.order_by(Trabajo.fecha_ingreso.desc()).all()
    
    # Listas para rellenar los componentes del buscador en el HTML
    clientes = Cliente.query.order_by(Cliente.empresa).all()
    maquinas = [t[0] for t in db.session.query(Trabajo.identificador)
                .filter(Trabajo.identificador != None, Trabajo.identificador != '')
                .distinct().order_by(Trabajo.identificador).all()]

    return render_template('historial.html', trabajos=lista,
                           cliente_q=cliente_q, maquina_q=maquina_q,
                           clientes=clientes, maquinas=maquinas)

# ═══════════════════════════════════════════════════════
# RUTAS — GASTOS GENERALES
# ═══════════════════════════════════════════════════════

@app.route('/gastos-generales', methods=['GET', 'POST'])
def gastos_generales():
    categorias = get_opciones('categoria_gasto')
    formas_pago = get_opciones('forma_pago')
    if request.method == 'POST':
        g = GastoGeneral(
            fecha=datetime.strptime(request.form['fecha'], '%Y-%m-%d').date(),
            categoria=request.form['categoria'],
            proveedor=request.form.get('proveedor'),
            monto=float(request.form['monto'] or 0),
            forma_pago=request.form.get('forma_pago'),
            comprobante=request.form.get('comprobante'),
            notas=request.form.get('notas')
        )
        db.session.add(g)
        db.session.commit()
        return redirect(url_for('gastos_generales'))
    lista = GastoGeneral.query.order_by(GastoGeneral.fecha.desc()).limit(50).all()
    return render_template('gastos_generales.html', gastos=lista,
                           categorias=categorias, formas_pago=formas_pago)

@app.route('/gastos-generales/<int:id>/anular', methods=['POST'])
def anular_gasto_general(id):
    """Anula un gasto general — queda registrado pero no se contabiliza"""
    g = GastoGeneral.query.get_or_404(id)
    g.anulado = True
    db.session.commit()
    return redirect(url_for('gastos_generales'))

# ═══════════════════════════════════════════════════════
# RUTAS — RESULTADOS
# ═══════════════════════════════════════════════════════

@app.route('/resultados')
def resultados():
    mes = int(request.args.get('mes', date.today().month))
    anio = int(request.args.get('anio', date.today().year))
    trabajos_mes = Trabajo.query.filter(
        db.extract('month', Trabajo.fecha_ingreso) == mes,
        db.extract('year', Trabajo.fecha_ingreso) == anio
    ).all()
    ventas = sum(t.presupuestado for t in trabajos_mes)
    cobros = Cobro.query.filter(
        db.extract('month', Cobro.fecha) == mes,
        db.extract('year', Cobro.fecha) == anio
    ).all()
    cobrado = sum(c.monto for c in cobros)
    saldo_total = sum(t.saldo for t in Trabajo.query.filter(Trabajo.estado != 'entregado').all())
    gastos_trabajos = GastoTrabajo.query.filter(
        db.extract('month', GastoTrabajo.fecha) == mes,
        db.extract('year', GastoTrabajo.fecha) == anio
    ).all()
    gastos_gen = GastoGeneral.query.filter(
    db.extract('month', GastoGeneral.fecha) == mes,
    db.extract('year', GastoGeneral.fecha) == anio,
    GastoGeneral.anulado == False
    ).all()
    cats = {}
    for g in gastos_trabajos + gastos_gen:
        cats[g.categoria] = cats.get(g.categoria, 0) + g.monto
    total_gastos = sum(cats.values())
    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    return render_template('resultados.html',
        mes=mes, anio=anio, meses=meses,
        ventas=ventas, cobrado=cobrado, saldo_total=saldo_total,
        total_gastos=total_gastos, cats=cats,
        resultado_financiero=cobrado - total_gastos,
        resultado_economico=ventas - total_gastos)

# ═══════════════════════════════════════════════════════
# RUTAS — CONFIGURACIÓN
# ═══════════════════════════════════════════════════════

@app.route('/configuracion')
def configuracion():
    """Pantalla para gestionar las opciones de los desplegables"""
    tipos = {
        'tipo_equipo': 'Tipos de equipo',
        'tipo_trabajo': 'Tipos de trabajo',
        'categoria_gasto': 'Categorías de gasto',
        'forma_pago': 'Formas de pago',
    }
    opciones = {}
    for tipo in tipos:
        opciones[tipo] = OpcionLista.query.filter_by(tipo=tipo).order_by(OpcionLista.orden, OpcionLista.valor).all()
    return render_template('configuracion.html', tipos=tipos, opciones=opciones)

@app.route('/configuracion/agregar', methods=['POST'])
def agregar_opcion():
    tipo = request.form['tipo']
    valor = request.form['valor'].strip()
    if valor:
        # Verificar que no exista ya
        existe = OpcionLista.query.filter_by(tipo=tipo, valor=valor).first()
        if not existe:
            db.session.add(OpcionLista(tipo=tipo, valor=valor))
            db.session.commit()
    return redirect(url_for('configuracion'))

@app.route('/configuracion/<int:id>/eliminar', methods=['POST'])
def eliminar_opcion(id):
    o = OpcionLista.query.get_or_404(id)
    db.session.delete(o)
    db.session.commit()
    return redirect(url_for('configuracion'))

# ═══════════════════════════════════════════════════════
# RUTAS — EXPORTAR EXCEL
# ═══════════════════════════════════════════════════════

@app.route('/exportar-excel')
def exportar_excel():
    """Genera un Excel con todas las tablas para usar en PowerBI"""
    wb = openpyxl.Workbook()

    navy = "1F3864"
    white = "FFFFFF"

    def hdr_style(cell, text):
        cell.value = text
        cell.font = Font(bold=True, color=white, name='Arial', size=10)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── Hoja Clientes ──────────────────────────────────
    ws = wb.active
    ws.title = "Clientes"
    headers = ['ID', 'Empresa', 'Contacto', 'Teléfono', 'Email', 'Dirección', 'CUIT']
    for ci, h in enumerate(headers, 1):
        hdr_style(ws.cell(1, ci), h)
    for c in Cliente.query.order_by(Cliente.empresa).all():
        ws.append([c.id, c.empresa, c.contacto, c.telefono, c.email, c.direccion, c.cuit])

    # ── Hoja Presupuestos ──────────────────────────────
    ws2 = wb.create_sheet("Presupuestos")
    headers2 = ['ID', 'Número', 'Cliente', 'Tipo Equipo', 'Identificador', 'Marca',
                'Modelo', 'Tipo Trabajo', 'Estado', 'Total', 'Fecha']
    for ci, h in enumerate(headers2, 1):
        hdr_style(ws2.cell(1, ci), h)
    for p in Presupuesto.query.order_by(Presupuesto.creado.desc()).all():
        ws2.append([p.id, p.numero, p.cliente.empresa, p.tipo_equipo, p.identificador,
                    p.marca, p.modelo, p.tipo_trabajo, p.estado, p.total,
                    p.creado.strftime('%d/%m/%Y')])

    # ── Hoja Trabajos ──────────────────────────────────
    ws3 = wb.create_sheet("Trabajos")
    headers3 = ['ID', 'Número', 'Cliente', 'Tipo Equipo', 'Identificador', 'Marca',
                'Modelo', 'Tipo Trabajo', 'Estado', 'Presupuestado', 'Cobrado',
                'Saldo', 'Gastos', 'Ganancia', 'F. Ingreso', 'F. Entrega']
    for ci, h in enumerate(headers3, 1):
        hdr_style(ws3.cell(1, ci), h)
    for t in Trabajo.query.order_by(Trabajo.creado.desc()).all():
        ws3.append([t.id, t.numero, t.cliente.empresa, t.tipo_equipo, t.identificador,
                    t.marca, t.modelo, t.tipo_trabajo, t.estado, t.presupuestado,
                    t.total_cobrado, t.saldo, t.total_gastos, t.ganancia,
                    t.fecha_ingreso.strftime('%d/%m/%Y') if t.fecha_ingreso else '',
                    t.fecha_entrega.strftime('%d/%m/%Y') if t.fecha_entrega else ''])

    # ── Hoja Gastos ────────────────────────────────────
    ws4 = wb.create_sheet("Gastos")
    headers4 = ['ID', 'Trabajo', 'Cliente', 'Fecha', 'Categoría', 'Proveedor', 'Descripción', 'Monto', 'Forma Pago']
    for ci, h in enumerate(headers4, 1):
        hdr_style(ws4.cell(1, ci), h)
    for g in GastoTrabajo.query.order_by(GastoTrabajo.fecha.desc()).all():
        ws4.append([g.id, g.trabajo.numero, g.trabajo.cliente.empresa,
                    g.fecha.strftime('%d/%m/%Y'), g.categoria, g.proveedor,
                    g.descripcion, g.monto, g.forma_pago])

    # ── Hoja Gastos Generales ──────────────────────────
    ws5 = wb.create_sheet("Gastos Generales")
    headers5 = ['ID', 'Fecha', 'Categoría', 'Proveedor', 'Monto', 'Forma Pago', 'Comprobante']
    for ci, h in enumerate(headers5, 1):
        hdr_style(ws5.cell(1, ci), h)
    for g in GastoGeneral.query.order_by(GastoGeneral.fecha.desc()).all():
        ws5.append([g.id, g.fecha.strftime('%d/%m/%Y'), g.categoria,
                    g.proveedor, g.monto, g.forma_pago, g.comprobante])

    # ── Hoja Cobranzas ─────────────────────────────────
    ws6 = wb.create_sheet("Cobranzas")
    headers6 = ['ID', 'Trabajo', 'Cliente', 'Fecha', 'Monto', 'Forma Pago', 'Comprobante', 'Notas']
    for ci, h in enumerate(headers6, 1):
        hdr_style(ws6.cell(1, ci), h)
    for c in Cobro.query.order_by(Cobro.fecha.desc()).all():
        ws6.append([c.id, c.trabajo.numero, c.trabajo.cliente.empresa,
                    c.fecha.strftime('%d/%m/%Y'), c.monto, c.forma_pago,
                    c.comprobante, c.notas])

    # Ajustar anchos de columna automáticamente
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f'TallerCosta_{date.today().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ═══════════════════════════════════════════════════════
# INICIO — Crear tablas y cargar datos por defecto
# ═══════════════════════════════════════════════════════

with app.app_context():
    db.create_all()       # Crea las tablas si no existen
    seed_opciones()       # Carga opciones por defecto si la tabla está vacía

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=8080, threaded=True)
